#!/usr/bin/env python3
from pathlib import Path
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MARKER_RE = re.compile(r'^(0\.|[1-5]\.[0-6]\.)')


def clean(v):
    if v is None:
        return ''
    s = str(v).strip()
    s = s.replace('\u00a0', ' ')
    s = re.sub(r'\s+', ' ', s)
    return s


def find_marker(row):
    for i, c in enumerate(row):
        if MARKER_RE.match(c):
            return i, c
    return None, ''


def pick(row, idx):
    return row[idx] if 0 <= idx < len(row) else ''


def map_p0_5(row, m):
    return [
        pick(row, m + 0),
        pick(row, m + 1),
        pick(row, m + 2),
        pick(row, m + 3),
        pick(row, m + 4),
        pick(row, m + 5),
        pick(row, m + 6),
        pick(row, m + 8),
        pick(row, m + 7),
        pick(row, m + 9),
        pick(row, m + 10),
        pick(row, m + 11),
        pick(row, m + 12),
        pick(row, m + 14),
        pick(row, m + 13),
        pick(row, m + 15),
        pick(row, m + 17),
    ]


def map_p6(row, m):
    return [
        pick(row, m + 0),
        pick(row, m + 1),
        pick(row, m + 2),
        pick(row, m + 3),
        pick(row, m + 4),
        pick(row, m + 5),
        pick(row, m + 6),
        pick(row, m + 8),
        pick(row, m + 7),
        pick(row, m + 9),
        pick(row, m + 10),
        pick(row, m + 11),
        pick(row, m + 12),
        pick(row, m + 14),
        pick(row, m + 13),
        pick(row, m + 15),
        pick(row, m + 16),
        pick(row, m + 17),
    ]


def apply_header_style(ws, cols):
    fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    font = Font(color='FFFFFF', bold=True)
    for c in range(1, cols + 1):
        ws.cell(1, c).fill = fill
        ws.cell(1, c).font = font
        ws.column_dimensions[get_column_letter(c)].width = 28
    ws.freeze_panes = 'A2'
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=cols):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')


def refresh(workbook_path):
    wb = load_workbook(workbook_path)
    if 'PDF_RAW_TABLE' not in wb.sheetnames:
        raise RuntimeError('PDF_RAW_TABLE sheet not found')

    raw = wb['PDF_RAW_TABLE']

    extracted = []
    for r in raw.iter_rows(min_row=1, max_row=raw.max_row, min_col=1, max_col=raw.max_column, values_only=True):
        row = [clean(v) for v in r]
        if any(row):
            extracted.append(row)

    buckets = {k: [] for k in ['Page-0','Page-1','Page-2','Page-3','Page-4','Page-5','Page-6']}
    marker_rows = []

    for i, row in enumerate(extracted, 1):
        mcol, marker = find_marker(row)
        if mcol is None:
            continue
        marker_rows.append((i, mcol, marker, row))
        if re.match(r'^0\.', marker):
            buckets['Page-0'].append((i, mcol, marker, row))
        if re.match(r'^[1-5]\.1\.', marker):
            buckets['Page-1'].append((i, mcol, marker, row))
        if re.match(r'^[1-5]\.2\.', marker):
            buckets['Page-2'].append((i, mcol, marker, row))
        if re.match(r'^[1-5]\.3\.', marker):
            buckets['Page-3'].append((i, mcol, marker, row))
        if re.match(r'^[1-5]\.4\.', marker):
            buckets['Page-4'].append((i, mcol, marker, row))
        if re.match(r'^[1-5]\.(5|6)\.', marker):
            buckets['Page-5'].append((i, mcol, marker, row))
        if re.match(r'^[1-5]\.[0-6]\.', marker):
            buckets['Page-6'].append((i, mcol, marker, row))

    # recreate index sheets and page sheets
    for name in ['ROW_MARKER_INDEX','PAGE_MAPPING_INDEX','Page-0','Page-1','Page-2','Page-3','Page-4','Page-5','Page-6']:
        if name in wb.sheetnames:
            wb.remove(wb[name])

    idx = wb.create_sheet('ROW_MARKER_INDEX')
    idx.append(['SOURCE_RAW_ROW','MARKER_COL_INDEX_0_BASED','SUBTOPIC_MARKER','RAW_COL_0','RAW_COL_1','RAW_COL_2','RAW_COL_3','RAW_COL_4'])
    for ridx,mcol,marker,row in marker_rows:
        idx.append([ridx,mcol,marker,pick(row,0),pick(row,1),pick(row,2),pick(row,3),pick(row,4)])
    apply_header_style(idx, 8)

    page_headers = [
        'ROW LABEL','What is it for? Why is it important? (Relevance)','What is it? (Introduction)',
        'How does it work successfully? (Effective Operating Procedure)',
        'What ultimately cause it to work successfully? (Ultimate Driving System)',
        'How do the Ultimate Driving System cause it to work successfully? (Success Mechanism)',
        'What principles are the Ultimate Driving System based on? (Effective Principle System)',
        'What tool(s) do the ultimate drivers require to work? (Ultimate Effective System)',
        'What environmental conditions do the ultimate drivers require to work? (Ultimate Effective System)',
        'How can it fail? (Failure Actions)',
        'What ultimately cause it to fail? (Ultimate Blocking System)',
        'How do the ultimate blockers cause it to fail? (Failure Mechanism)',
        'What principles are the ultimate blockers based on? (Risky Principles)',
        'What tool(s) do the ultimate blockers require to work? (Risky Tools)',
        'What environmental conditions do the ultimate blockers require to work? (Risky Environments)',
        'What to do if it fails? (What else?)',
        'Next Steps to Take (Now What? Now How?)',
        'SOURCE_RAW_ROW','MARKER_COL_INDEX_0_BASED'
    ]

    page6_headers = [
        'SUB-TOPIC X.0–X.6','What is it for? Why is it important? (Relevance)','What is it? (Introduction)',
        'How does it work successfully? (Success Actions)',
        'What ultimately cause it to work successfully? (Ultimate Drivers)',
        'How do the Ultimate Driving System cause it to work successfully? (Success Mechanism)',
        'What principles are the Ultimate Driving System based on? (Effective Principle System)',
        'What tool(s) do the ultimate drivers require to work? (Ultimate Effective System)',
        'What environmental conditions do the ultimate drivers require to work? (Ultimate Effective System)',
        'How can it fail? (Failure Actions)',
        'What ultimately cause it to fail? (Ultimate Blockers)',
        'How do the ultimate blockers cause it to fail? (Failure Mechanism)',
        'What principles are the ultimate blockers based on? (Failure Principles)',
        'What tool(s) do the ultimate blockers require to work? (Failure Tools)',
        'What environmental conditions do the ultimate blockers require to work? (Failure Environments)',
        'What to do if it fails? (What else?)','Other Questions (Others)',
        'Next Steps to Take (Now What? Now How?)',
        'SOURCE_RAW_ROW','MARKER_COL_INDEX_0_BASED'
    ]

    for page in ['Page-0','Page-1','Page-2','Page-3','Page-4','Page-5']:
        ws = wb.create_sheet(page)
        ws.append(page_headers)
        for ridx,mcol,marker,row in buckets[page]:
            ws.append(map_p0_5(row, mcol) + [ridx,mcol])
        apply_header_style(ws, len(page_headers))

    ws6 = wb.create_sheet('Page-6')
    ws6.append(page6_headers)
    for ridx,mcol,marker,row in buckets['Page-6']:
        ws6.append(map_p6(row, mcol) + [ridx,mcol])
    apply_header_style(ws6, len(page6_headers))

    pidx = wb.create_sheet('PAGE_MAPPING_INDEX')
    pidx.append(['PAGE','FILTER_RULE_ON_MARKER','ROW_COUNT'])
    rules = {
        'Page-0': r'^0\.',
        'Page-1': r'^[1-5]\.1\.',
        'Page-2': r'^[1-5]\.2\.',
        'Page-3': r'^[1-5]\.3\.',
        'Page-4': r'^[1-5]\.4\.',
        'Page-5': r'^[1-5]\.(5|6)\.',
        'Page-6': r'^[1-5]\.[0-6]\.'
    }
    for p in ['Page-0','Page-1','Page-2','Page-3','Page-4','Page-5','Page-6']:
        pidx.append([p, rules[p], len(buckets[p])])
    apply_header_style(pidx, 3)

    wb.save(workbook_path)

    return {k: len(v) for k,v in buckets.items()}


if __name__ == '__main__':
    import sys
    if len(sys.argv) != 2:
        print('Usage: python scripts/refresh_distilled_pages_from_raw_xlsx.py <xlsx_path>')
        raise SystemExit(1)
    path = Path(sys.argv[1])
    counts = refresh(path)
    print('Refreshed:', path)
    for k in ['Page-0','Page-1','Page-2','Page-3','Page-4','Page-5','Page-6']:
        print(k, counts[k])
